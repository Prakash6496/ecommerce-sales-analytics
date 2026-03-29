import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from analyze import run_analysis
from pathlib import Path

def style_header_row(ws, row_num, num_cols, bg_color="2E86AB"):
    """Apply header styling to a row"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=bg_color, 
                               end_color=bg_color, 
                               fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

def style_data_rows(ws, start_row, end_row, num_cols):
    """Apply alternating row colors"""
    for row in range(start_row, end_row + 1):
        fill_color = "F0F8FF" if row % 2 == 0 else "FFFFFF"
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=fill_color,
                                   end_color=fill_color,
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

def auto_fit_columns(ws):
    """Auto fit column widths"""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 4

def write_dataframe(ws, df, start_row, title, bg_color="2E86AB"):
    """Write a dataframe to worksheet with styling"""
    # Write title
    ws.cell(row=start_row, column=1, value=title)
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=13)
    ws.cell(row=start_row, column=1).fill = PatternFill(
        start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    ws.cell(row=start_row, column=1).font = Font(
        bold=True, size=13, color="FFFFFF")
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=len(df.columns))

    # Write headers
    header_row = start_row + 1
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=header_row, column=col_idx, 
                value=col_name.replace("_", " ").title())
    style_header_row(ws, header_row, len(df.columns), bg_color)

    # Write data
    for row_idx, row in enumerate(df.itertuples(index=False), 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=header_row + row_idx, 
                   column=col_idx, value=value)
    style_data_rows(ws, header_row + 1, 
                   header_row + len(df), len(df.columns))

    return header_row + len(df) + 2  # return next available row

def generate_report():
    print("\n📝 Generating Excel Report...")

    # Run all analysis queries
    results = run_analysis()

    # Create workbook
    wb = Workbook()

    # ─────────────────────────────────────
    # Sheet 1 — Summary Dashboard
    # ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "📊 Summary Dashboard"

    # Title
    ws1["A1"] = "🛒 E-Commerce Sales Analytics Report"
    ws1["A1"].font = Font(bold=True, size=16, color="1B4F72")
    ws1["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1["A2"].font = Font(italic=True, size=11, color="666666")
    ws1["A3"] = "Dataset: Olist Brazilian E-Commerce | Period: 2016-2018"
    ws1["A3"].font = Font(italic=True, size=11, color="666666")

    # Key Metrics
    ws1["A5"] = "📈 Key Business Metrics"
    ws1["A5"].font = Font(bold=True, size=13, color="1B4F72")

    metrics = [
        ("Total Orders",          "99,441"),
        ("Total Records Analyzed","551,535+"),
        ("Top Category",          "Health & Beauty"),
        ("Top City",              "São Paulo"),
        ("Top Payment Method",    "Credit Card (74%)"),
        ("Total Revenue",         "R$ 15,422,461"),
        ("High Value Orders",     "4,296 orders > R$500"),
        ("Max Single Order",      "R$ 13,664.08"),
    ]

    ws1["A6"] = "Metric"
    ws1["B6"] = "Value"
    style_header_row(ws1, 6, 2)

    for idx, (metric, value) in enumerate(metrics, 7):
        ws1.cell(row=idx, column=1, value=metric)
        ws1.cell(row=idx, column=2, value=value)
    style_data_rows(ws1, 7, 14, 2)
    auto_fit_columns(ws1)

    # ─────────────────────────────────────
    # Sheet 2 — Top Categories
    # ─────────────────────────────────────
    ws2 = wb.create_sheet("🏆 Top Categories")
    write_dataframe(ws2, results["top_categories"],
                   1, "Top 10 Revenue Generating Categories")
    auto_fit_columns(ws2)

    # ─────────────────────────────────────
    # Sheet 3 — Monthly Trends
    # ─────────────────────────────────────
    ws3 = wb.create_sheet("📅 Monthly Trends")
    write_dataframe(ws3, results["monthly_trends"],
                   1, "Monthly Sales Trends (2016-2018)")
    auto_fit_columns(ws3)

    # ─────────────────────────────────────
    # Sheet 4 — Regional Analysis
    # ─────────────────────────────────────
    ws4 = wb.create_sheet("🗺️ Regional Analysis")
    next_row = write_dataframe(ws4, results["top_cities"],
                              1, "Top 10 Cities by Revenue")
    write_dataframe(ws4, results["regional_revenue"],
                   next_row, "Top 10 States by Revenue")
    auto_fit_columns(ws4)

    # ─────────────────────────────────────
    # Sheet 5 — Payment Analysis
    # ─────────────────────────────────────
    ws5 = wb.create_sheet("💳 Payment Analysis")
    write_dataframe(ws5, results["payment_methods"],
                   1, "Payment Method Analysis")
    auto_fit_columns(ws5)

    # ─────────────────────────────────────
    # Sheet 6 — Advanced SQL Results
    # ─────────────────────────────────────
    ws6 = wb.create_sheet("🔍 Advanced SQL")
    next_row = write_dataframe(ws6, results["high_value_orders"],
                              1, "High Value Orders (CTE Query)")
    next_row = write_dataframe(ws6, results["running_total"],
                              next_row, "Running Total Revenue (Window Function)")
    write_dataframe(ws6, results["above_avg_states"],
                   next_row, "Above Average States (Subquery)")
    auto_fit_columns(ws6)

    # ─────────────────────────────────────
    # Save Report
    # ─────────────────────────────────────
    output_path = Path(__file__).parent / "reports"
    output_path.mkdir(exist_ok=True)

    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = output_path / filename

    wb.save(filepath)
    print(f"\n✅ Report saved to: {filepath}")
    print(f"📊 Report contains 6 sheets with full analysis!")
    return filepath

if __name__ == "__main__":
    generate_report()